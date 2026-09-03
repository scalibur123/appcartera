#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
reconstruir_historico.py

Reconstruye dia a dia, desde el 2 de enero de 2026, la serie de:

    latente(F) = Σ(titulos × cierre(F) en €) − Σ coste     de lo abierto ese dia

y con ella calcula, segun la formula acordada:

    GANANCIA(periodo) = latente(hoy) − latente(inicio)
                      + plusvalias BRUTAS de ventas del periodo   (col Y)
                      + dividendos BRUTOS del periodo             (col H)

NO SUBE NADA A SUPABASE. Solo calcula, valida contra el propio Excel
y deja el resultado en historico_latentes.csv para revisarlo.

La validacion es la clave: la latente que calcula para HOY tiene que
coincidir con la suma de la columna AB del Excel. Si no coincide,
el metodo esta mal y hay que pararse.

Uso:  python3 reconstruir_historico.py
"""

import csv
import json
import re
import sys
import time
import urllib.parse
import urllib.request
from datetime import datetime, timedelta, date
from pathlib import Path

import openpyxl

# ── Configuracion ─────────────────────────────────────────────────────
EXCEL = Path.home() / "Library/Mobile Documents/com~apple~CloudDocs/INVERSION/PLUSVALIAS BOLSA 26_sinmacros_Claude.xlsm"
DIR = Path(__file__).resolve().parent
TICKERS_JSON = DIR / "tickers.json"
OVERRIDE_JSON = DIR / "tickers_override.json"
CACHE_SYMS = DIR / "historico_symbols.json"
SALIDA_CSV = DIR / "historico_latentes.csv"
SALIDA_JSON = DIR / "serie_latentes.json"
# Dias pasados YA calculados con datos completos. Se escriben una vez y no se
# vuelven a tocar: si manana Yahoo sirve otra cosa, el pasado no se mueve.
CONGELADO = DIR / "historico_congelado.json"
# Cierres historicos ya descargados. Un cierre pasado no cambia nunca, asi que
# se acumulan en disco. Yahoo trunca al azar unos pocos .MC en cada ejecucion;
# con esta cache, un dia bajado bien queda guardado para siempre.
CACHE_CIERRES = DIR / "cierres_historicos.json"
# Fecha del maximo de cada simbolo dentro del historico que tenemos guardado.
SALIDA_MAXIMOS = DIR / "maximos.json"

HOJA = "2026"
HOJA_DIV = "dividendos 26"
ANIO = 2026
DESDE = date(ANIO, 1, 1)

MARCA_INI = "ejecuciones"
MARCA_FIN = "final app cartera diseno"

UA = {"User-Agent": "Mozilla/5.0"}

# Wall Street cierra a las 22:00 hora espanola (21:00 en las pocas semanas de
# desfase entre el cambio de hora de EEUU y el de Europa). A las 22:10 la sesion
# esta cerrada SIEMPRE.
CIERRE_USA_LOCAL = (22, 10)


def _usa_ya_cerro(ahora=None):
    """True si la sesion de Wall Street de HOY ya ha cerrado."""
    ahora = ahora or datetime.now()
    return (ahora.hour, ahora.minute) >= CIERRE_USA_LOCAL


def _dia_provisional(f, hay_usd):
    """True si el dia f no se puede dar por cerrado todavia: es hoy, hay
    posiciones en dolares y Wall Street sigue abierta. Los precios USA que
    devuelve Yahoo son entonces intradia, no cierres, y publicar ese dia hace
    que ESTA SEMANA reste contra una base falsa."""
    if not hay_usd:
        return False
    return f == date.today() and not _usa_ya_cerro()



def sin_acentos(s):
    rep = {'á':'a','é':'e','í':'i','ó':'o','ú':'u','ü':'u','ñ':'n',
           'Á':'a','É':'e','Í':'i','Ó':'o','Ú':'u','Ü':'u','Ñ':'n'}
    t = "".join(rep.get(c, c) for c in str(s)).lower()
    return " ".join(t.split())


def barra(t):
    print("\n" + "=" * 74)
    print(t)
    print("=" * 74)


def fmt(v, d=2):
    return f"{v:,.{d}f}".replace(",", "\x00").replace(".", ",").replace("\x00", ".")


# ══════════════════════════════════════════════════════════════════════
# 1. LEER EL EXCEL
# ══════════════════════════════════════════════════════════════════════
def leer_excel():
    if not EXCEL.exists():
        sys.exit(f"❌ Excel no encontrado: {EXCEL}")
    wb = openpyxl.load_workbook(open(str(EXCEL), "rb"), read_only=True, data_only=True)
    ws = wb[HOJA]

    filas = {}
    ini = fin = None
    for r, row in enumerate(ws.iter_rows(min_row=1, max_row=1500, max_col=32, values_only=True), 1):
        filas[r] = row
        for v in row:
            if isinstance(v, str):
                t = sin_acentos(v)
                if MARCA_INI in t and ini is None and r >= 50:
                    ini = r
                if MARCA_FIN in t and ini and r > ini and fin is None:
                    fin = r
    if not ini:
        sys.exit("❌ No se encuentra 'ejecuciones' en la hoja 2026")

    def cel(r, c):
        row = filas.get(r)
        return row[c - 1] if row and len(row) >= c else None

    pos = (5, ini - 1)
    ven = (ini + 1, (fin - 1) if fin else 1500)
    print(f"📐 POSICIONES {pos[0]}-{pos[1]} · VENTAS {ven[0]}-{ven[1]}")

    posiciones = []   # cada una: abierta desde f_compra hasta f_venta (o None)
    latente_excel = 0.0

    # Bloque de posiciones ABIERTAS
    for r in range(pos[0], pos[1] + 1):
        t = cel(r, 4)
        if not t or not str(t).strip():
            continue
        K, N = cel(r, 11), cel(r, 14)
        if not isinstance(K, (int, float)) or not isinstance(N, (int, float)):
            continue
        f_compra_raw = cel(r, 9)
        if isinstance(f_compra_raw, datetime):
            f_compra, fecha_ok = f_compra_raw.date(), True
        else:
            f_compra, fecha_ok = date(2000, 1, 1), False
        posiciones.append({
            "fila": r, "tckr": str(t).strip(), "moneda": str(cel(r, 7) or "EUR").strip().upper(),
            "titulos": float(K), "coste": float(N),
            "f_compra": f_compra, "f_venta": None, "fecha_ok": fecha_ok,
        })
        AB = cel(r, 28)
        if isinstance(AB, (int, float)):
            latente_excel += AB

    # Bloque de VENTAS (posiciones cerradas este año)
    ventas = []
    for r in range(ven[0], ven[1] + 1):
        t = cel(r, 4)
        q = cel(r, 17)
        if not t or not isinstance(q, datetime):
            continue
        K, N = cel(r, 11), cel(r, 14)
        Y = cel(r, 25)
        if not isinstance(K, (int, float)) or not isinstance(N, (int, float)):
            continue
        f_compra_raw = cel(r, 9)
        if isinstance(f_compra_raw, datetime):
            f_compra, fecha_ok = f_compra_raw.date(), True
        else:
            f_compra, fecha_ok = date(2000, 1, 1), False
        posiciones.append({
            "fila": r, "tckr": str(t).strip(), "moneda": str(cel(r, 7) or "EUR").strip().upper(),
            "titulos": float(K), "coste": float(N),
            "f_compra": f_compra, "f_venta": q.date(), "fecha_ok": fecha_ok,
        })
        if isinstance(Y, (int, float)):
            ventas.append({"fecha": q.date(), "tckr": str(t).strip(), "bruto": float(Y)})

    # Dividendos: col A fecha, col H Total Bruto (ya en euros)
    dividendos = []
    if HOJA_DIV in wb.sheetnames:
        for row in wb[HOJA_DIV].iter_rows(min_row=4, max_row=500, max_col=10, values_only=True):
            f, h = row[0], row[7]
            if isinstance(f, datetime) and isinstance(h, (int, float)) and f.year == ANIO:
                dividendos.append({"fecha": f.date(), "bruto": float(h)})
    wb.close()

    abiertas = [p for p in posiciones if p["f_venta"] is None]
    print(f"   {len(abiertas)} posiciones abiertas · {len(ventas)} ventas · {len(dividendos)} dividendos")
    coste_excel = sum(p["coste"] for p in posiciones if p["f_venta"] is None)
    print(f"   Latente HOY segun columna AB del Excel: {fmt(latente_excel)} €")
    print(f"   Coste  HOY segun columna N  del Excel: {fmt(coste_excel)} €")

    # ── FECHAS DE COMPRA NO LEIDAS ────────────────────────────────────
    # Una posicion sin fecha de compra legible se trataba como comprada en el
    # año 2000: entraba en la serie desde el primer dia, con el coste completo
    # pagado meses despues. Eso reescribe el pasado en cada ejecucion.
    rotas = [p for p in posiciones if not p.get("fecha_ok")]
    if rotas:
        barra("⚠️  FECHAS DE COMPRA NO LEIDAS (columna J)")
        coste_roto = sum(p["coste"] for p in rotas)
        for p in sorted(rotas, key=lambda x: -x["coste"]):
            estado = "vendida " + str(p["f_venta"]) if p["f_venta"] else "abierta"
            print(f"   fila {p['fila']:>4}  {p['tckr']:<10} coste {fmt(p['coste']):>14} €   {estado}")
        pct = (coste_roto / coste_excel * 100) if coste_excel else 0.0
        print(f"\n   {len(rotas)} posiciones sin fecha legible")
        print(f"   Coste afectado: {fmt(coste_roto)} € ({pct:.1f}% de la cartera)")
        print("   Estas posiciones falsean TODOS los dias pasados de la serie.")
    else:
        print("   ✅ Todas las fechas de compra se han leido correctamente")

    return posiciones, ventas, dividendos, latente_excel, coste_excel


# ══════════════════════════════════════════════════════════════════════
# 2. RESOLVER SIMBOLOS DE YAHOO
# ══════════════════════════════════════════════════════════════════════
def buscar_en_yahoo(tckr, moneda):
    url = ("https://query2.finance.yahoo.com/v1/finance/search?q="
           + urllib.parse.quote(tckr) + "&quotesCount=8&newsCount=0")
    try:
        req = urllib.request.Request(url, headers=UA)
        d = json.loads(urllib.request.urlopen(req, timeout=10).read().decode())
    except Exception:
        return None
    cands = [q for q in d.get("quotes", []) if q.get("symbol")]
    if not cands:
        return None
    if moneda == "USD":
        for q in cands:
            s = q["symbol"]
            if "." not in s and q.get("exchange") in ("NMS", "NYQ", "NGM", "ASE", "PCX", "BTS"):
                return s
    for q in cands:
        s = q["symbol"]
        if s.upper().startswith(tckr.upper()):
            return s
    return cands[0]["symbol"]


def resolver_simbolos(posiciones):
    barra("RESOLVIENDO SIMBOLOS DE YAHOO")
    mapa = {}
    for f in (TICKERS_JSON, OVERRIDE_JSON, CACHE_SYMS):
        if f.exists():
            try:
                d = json.loads(f.read_text(encoding="utf-8"))
                for k, v in d.items():
                    mapa[k] = v.get("symbol") if isinstance(v, dict) else v
            except Exception:
                pass

    necesarios = {}
    for p in posiciones:
        necesarios.setdefault(p["tckr"], p["moneda"])

    nuevos, fallidos = {}, []
    for tckr, moneda in sorted(necesarios.items()):
        if mapa.get(tckr):
            continue
        if moneda == "USD":
            mapa[tckr] = tckr
            nuevos[tckr] = tckr
            continue
        s = buscar_en_yahoo(tckr, moneda)
        time.sleep(0.3)
        if s:
            mapa[tckr] = s
            nuevos[tckr] = s
        else:
            fallidos.append(tckr)

    if nuevos:
        cache = {}
        if CACHE_SYMS.exists():
            try:
                cache = json.loads(CACHE_SYMS.read_text(encoding="utf-8"))
            except Exception:
                pass
        cache.update(nuevos)
        CACHE_SYMS.write_text(json.dumps(cache, indent=2, ensure_ascii=False), encoding="utf-8")
        print(f"  {len(nuevos)} simbolos nuevos resueltos y guardados en {CACHE_SYMS.name}")
        for k, v in sorted(nuevos.items()):
            print(f"    {k:<8} -> {v}")
    if fallidos:
        print(f"  ⚠️  {len(fallidos)} SIN RESOLVER: {', '.join(fallidos)}")
        print(f"      Añadelos a mano a {CACHE_SYMS.name} y vuelve a ejecutar.")
    return mapa, fallidos


# ══════════════════════════════════════════════════════════════════════
# 3. DESCARGAR CIERRES HISTORICOS
# ══════════════════════════════════════════════════════════════════════
def _pedir(url):
    req = urllib.request.Request(url, headers=UA)
    d = json.loads(urllib.request.urlopen(req, timeout=20).read().decode())
    res = d["chart"]["result"][0]
    ts = res["timestamp"]
    cl = res["indicators"]["quote"][0]["close"]
    return {datetime.utcfromtimestamp(t).date(): c
            for t, c in zip(ts, cl) if c is not None}


def descargar(simbolo, desde):
    """Cierres diarios desde 'desde'.

    Yahoo, cuando estrangula por exceso de peticiones, IGNORA period1/period2 y
    responde con su rango por defecto (~1 mes) sin dar ningun error. El script
    antiguo aceptaba esa respuesta corta como buena y los dias anteriores se
    calculaban sin esa posicion. Aqui se comprueba que el histórico devuelto
    llegue de verdad hasta 'desde', y si no, se reintenta esperando mas.
    """
    tope = desde - timedelta(days=20)
    p1 = int(datetime.combine(tope, datetime.min.time()).timestamp())
    p2 = int(time.time())
    base = f"https://query1.finance.yahoo.com/v8/finance/chart/{urllib.parse.quote(simbolo)}"
    intentos = [
        f"{base}?period1={p1}&period2={p2}&interval=1d",
        f"{base}?range=2y&interval=1d",
        f"{base}?period1={p1}&period2={p2}&interval=1d",
        f"{base}?range=2y&interval=1d",
    ]
    mejor = {}
    for n, url in enumerate(intentos):
        try:
            serie = _pedir(url)
        except Exception:
            serie = {}
        if len(serie) > len(mejor):
            mejor = serie
        if mejor and min(mejor) <= desde:
            return mejor          # cobertura suficiente, no hace falta insistir
        time.sleep(1.5 * (n + 1))  # espera creciente: es estrangulamiento
    return mejor


def _cargar_cache():
    try:
        crudo = json.loads(CACHE_CIERRES.read_text(encoding="utf-8")) if CACHE_CIERRES.exists() else {}
    except Exception:
        crudo = {}
    return {sim: {date.fromisoformat(f): v for f, v in d.items()} for sim, d in crudo.items()}


def _guardar_cache(cache):
    crudo = {sim: {str(f): round(v, 6) for f, v in d.items()} for sim, d in cache.items()}
    CACHE_CIERRES.write_text(json.dumps(crudo, ensure_ascii=False, sort_keys=True), encoding="utf-8")


def descargar_todo(simbolos):
    barra("DESCARGANDO CIERRES HISTORICOS DE YAHOO")
    cache = _cargar_cache()
    print(f"  Cache en disco: {len(cache)} simbolos")
    series, vacios = {}, []
    rescatados, splits = [], []
    total = len(simbolos)
    for n, s in enumerate(sorted(simbolos), 1):
        nueva = descargar(s, DESDE)
        guardada = cache.get(s, {})

        # ¿Split? Si en fechas comunes los precios difieren mucho, la serie
        # guardada esta sin ajustar y hay que tirarla.
        if nueva and guardada:
            comunes = [f for f in nueva if f in guardada][:20]
            if comunes:
                dif = sum(abs(nueva[f] - guardada[f]) / max(abs(guardada[f]), 1e-9) for f in comunes) / len(comunes)
                if dif > 0.05:
                    splits.append(s)
                    guardada = {}

        fusion = dict(guardada)
        fusion.update(nueva)
        if fusion:
            if nueva and guardada and min(nueva) > min(guardada):
                rescatados.append(s)
            series[s] = fusion
            cache[s] = fusion
        else:
            vacios.append(s)
        if n % 25 == 0:
            print(f"    {n}/{total}...")
        time.sleep(0.25)

    _guardar_cache(cache)
    if splits:
        print(f"  ↺ serie guardada descartada por ajuste/split: {', '.join(splits)}")
    if rescatados:
        print(f"  💾 {len(rescatados)} simbolos truncados hoy, rescatados de la cache:")
        print(f"     {', '.join(sorted(rescatados))}")
    print(f"  ✅ {len(series)}/{total} con datos")
    if vacios:
        print(f"  ❌ sin datos: {', '.join(vacios)}")

    # Simbolos cuyo historico NO llega al inicio del periodo. Sin esto, los dias
    # anteriores a su primer cierre se calculan sin esa posicion (ni su valor ni
    # su coste) y la latente de esas fechas sale falseada.
    cortos = []
    for sim, serie in series.items():
        if not serie:
            continue
        primera = min(serie)
        if primera > DESDE:
            cortos.append((sim, primera, max(serie), len(serie)))
    if cortos:
        barra("⚠️  SIMBOLOS SIN HISTORICO COMPLETO")
        for sim, pri, ult, n in sorted(cortos, key=lambda x: x[1], reverse=True):
            print(f"   {sim:<12} solo desde {pri}  (hasta {ult}, {n} cierres)")
        print(f"\n   {len(cortos)} simbolos con historico truncado.")
    return series, vacios


def valor_en(serie, f):
    """Ultimo cierre disponible en o antes de f (arrastra festivos)."""
    if f in serie:
        return serie[f]
    cand = [d for d in serie if d <= f]
    return serie[max(cand)] if cand else None


# ══════════════════════════════════════════════════════════════════════
# 4. RECONSTRUIR LA SERIE
# ══════════════════════════════════════════════════════════════════════
def _habil_antes(d):
    d = d - timedelta(days=1)
    while d.weekday() >= 5:
        d = d - timedelta(days=1)
    return d


def _salud(serie_out, hoy):
    """Estado de las tres bases que usa la app, para que pueda avisar."""
    pub = sorted(str(x["fecha"]) for x in serie_out
                 if not x["incompleto"] and not x.get("provisional"))
    lunes = hoy - timedelta(days=hoy.weekday())
    dia1 = hoy.replace(day=1)
    ene1 = hoy.replace(month=1, day=1)
    out = {}
    for nombre, ini in (("semana", lunes), ("mes", dia1), ("anual", ene1)):
        ini_s = ini.strftime("%Y-%m-%d")
        previas = [f for f in pub if f < ini_s]
        esperada = _habil_antes(ini).strftime("%Y-%m-%d")
        if not previas:
            out[nombre] = {"ok": False, "base": None, "esperada": esperada,
                           "motivo": "no hay ningun dia anterior al periodo"}
        else:
            base = previas[-1]
            ok = base >= esperada
            out[nombre] = {"ok": ok, "base": base, "esperada": esperada,
                           "motivo": "" if ok else "la base va con retraso"}
    out["ultimo_dia"] = pub[-1] if pub else None
    return out


def main():
    posiciones, ventas, dividendos, latente_excel, coste_excel = leer_excel()
    mapa, fallidos = resolver_simbolos(posiciones)

    simbolos = {mapa[p["tckr"]] for p in posiciones if mapa.get(p["tckr"])}
    simbolos.add("EURUSD=X")
    series, vacios = descargar_todo(simbolos)

    fx = series.get("EURUSD=X", {})
    if not fx:
        sys.exit("❌ No hay serie de EUR/USD. Sin eso no se puede convertir nada.")

    hoy = date.today()
    # arrancamos unos dias antes del 1 de enero para tener base del ANUAL
    arranque = DESDE - timedelta(days=25)
    # El calendario sale de las series de ACCIONES, nunca del EUR/USD: las divisas
    # cotizan en fin de semana y con horarios que desplazan las fechas, y eso metia
    # domingos como base de la semana. Solo lunes-viernes, y solo dias en los que
    # cotiza una parte significativa de la cartera (descarta festivos sueltos).
    cuenta = {}
    for sim, serie in series.items():
        if sim == "EURUSD=X":
            continue
        for d in serie:
            if d.weekday() < 5:
                cuenta[d] = cuenta.get(d, 0) + 1
    if not cuenta:
        sys.exit("❌ No hay series de acciones para construir el calendario.")
    umbral = max(cuenta.values()) * 0.30
    dias = sorted(d for d, n in cuenta.items() if n >= umbral and arranque <= d <= hoy)
    if not dias:
        sys.exit("❌ No hay dias habiles en el rango.")
    print(f"  Calendario: {len(dias)} dias habiles de {dias[0]} a {dias[-1]}")

    barra("RECONSTRUYENDO LA SERIE DIARIA")
    serie_out = []
    sin_precio_hoy = []
    for f in dias:
        cambio = valor_en(fx, f) or 1.0
        valor = coste = 0.0
        n_pos = 0
        # Coste de posiciones ABIERTAS ese dia para las que no hay precio.
        # Cuando falta el precio la posicion se cae entera: no suma valor pero
        # TAMPOCO suma coste. La composicion del dia cambia segun lo que Yahoo
        # devuelva en cada ejecucion, y por eso la latente de una fecha pasada
        # sale distinta cada vez. Esto lo mide.
        falta_coste = 0.0
        falta_tckrs = []
        n_usd = 0
        for p in posiciones:
            if p["f_compra"] > f:
                continue
            if p["f_venta"] is not None and p["f_venta"] <= f:
                continue
            sim = mapa.get(p["tckr"])
            s = series.get(sim) if sim else None
            px = valor_en(s, f) if s else None
            if px is None:
                falta_coste += p["coste"]
                falta_tckrs.append(p["tckr"])
                if f == dias[-1]:
                    sin_precio_hoy.append(p["tckr"])
                continue
            pe = px / cambio if p["moneda"] == "USD" else px
            valor += p["titulos"] * pe
            coste += p["coste"]
            n_pos += 1
            if p["moneda"] == "USD":
                n_usd += 1
        coste_abierto = coste + falta_coste
        # Un dia al que le falta coste no es comparable con un dia completo:
        # la resta de latentes mide el hueco, no el mercado. Se marca y NO se
        # publica en el JSON, para que la app diga "sin base" en vez de mentir.
        incompleto = falta_coste > 0.005 * coste_abierto if coste_abierto else True
        serie_out.append({"fecha": f, "valor": valor, "coste": coste, "latente": valor - coste,
                          "n_pos": n_pos, "n_usd": n_usd,
                          "falta_coste": falta_coste, "falta_tckrs": falta_tckrs,
                          "coste_abierto": coste_abierto, "incompleto": incompleto,
                          "provisional": _dia_provisional(f, n_usd > 0)})

    # ── HUELLA DE LAS FECHAS BASE ─────────────────────────────────────
    # Si el coste de una fecha base cambia entre ejecuciones, la serie no es
    # reproducible y el ANUAL/MES de la app se movera solo.
    # ── CONGELAR EL PASADO ────────────────────────────────────────────
    # Un dia completo y ya cerrado se guarda para siempre. En ejecuciones
    # posteriores se REUTILIZA el valor guardado en vez de recalcularlo.
    try:
        congelado = json.loads(CONGELADO.read_text(encoding="utf-8")) if CONGELADO.exists() else {}
    except Exception:
        congelado = {}
    hoy_str = str(dias[-1])
    nuevos, reutilizados, derivas = 0, 0, []
    for x in serie_out:
        f = str(x["fecha"])
        if f in congelado:
            g = congelado[f]
            if not x["incompleto"] and abs(x["latente"] - g["latente"]) > max(500, abs(g["latente"]) * 0.005):
                derivas.append((f, g["latente"], x["latente"]))
            x["valor"], x["coste"], x["latente"] = g["valor"], g["coste"], g["latente"]
            x["incompleto"] = False
            x["congelado"] = True
            reutilizados += 1
        elif not x["incompleto"] and not x["provisional"] and f < hoy_str:
            congelado[f] = {"valor": round(x["valor"], 2), "coste": round(x["coste"], 2),
                            "latente": round(x["latente"], 2), "n_pos": x["n_pos"],
                            "grabado": datetime.now().isoformat(timespec="seconds")}
            x["congelado"] = True
            nuevos += 1
        else:
            x["congelado"] = False
    CONGELADO.write_text(json.dumps(congelado, indent=1, ensure_ascii=False, sort_keys=True), encoding="utf-8")
    # ── FECHA DEL MAXIMO: YA NO SE ESCRIBE AQUI ───────────────────────
    # Estas series son de CIERRES y solo arrancan en la fecha de inicio de
    # la reconstruccion, asi que daban maximos falsos (ZS: 242,08 desde
    # diciembre cuando el real eran 336,99 del 3-nov-2025). maximos.json lo
    # genera ahora actualizar_maximos.py, que lee de Yahoo los maximos
    # intradia de 52 semanas: el mismo dato que muestra la app.
    print("ℹ️  maximos.json no se toca aqui (lo genera actualizar_maximos.py)")

    barra("HISTORICO CONGELADO")
    print(f"   {reutilizados} dias reutilizados del archivo (no recalculados)")
    print(f"   {nuevos} dias nuevos congelados")
    print(f"   {len(congelado)} dias en total en {CONGELADO.name}")
    if derivas:
        print(f"\n   ⚠️  {len(derivas)} dias en los que Yahoo daria hoy algo distinto")
        print("      (se mantiene el valor congelado, que es el bueno):")
        for f, viejo, nuevo in derivas[:10]:
            print(f"      {f}  congelado {fmt(viejo)} vs ahora {fmt(nuevo)}")

    malos = [x for x in serie_out if x["incompleto"]]
    if malos:
        barra("⚠️  DIAS DESCARTADOS POR COSTE SIN PRECIO")
        print(f"   {len(malos)} de {len(serie_out)} dias no se publican en la serie.")
        print(f"   Del {malos[0]['fecha']} al {malos[-1]['fecha']}")
        peor = max(malos, key=lambda x: x["falta_coste"])
        print(f"   Peor dia: {peor['fecha']} con {fmt(peor['falta_coste'])} € de coste sin precio")
        print("   Mientras existan estos huecos, MES y ANUAL saldran 'sin base' en la app.")
        print("   Es lo correcto: mejor sin dato que con un dato inventado.")

    barra("HUELLA DE LAS FECHAS BASE (composicion del dia)")
    por_fecha = {str(x["fecha"]): x for x in serie_out}
    claves = [k for k in sorted(por_fecha) if k < f"{ANIO}-01-01"][-1:]
    claves += [k for k in sorted(por_fecha) if k.startswith(f"{ANIO}-07")][-1:]
    claves += [k for k in sorted(por_fecha)][-6:]
    vistos = set()
    for k in claves:
        if k in vistos:
            continue
        vistos.add(k)
        d = por_fecha[k]
        print(f"\n  {k}   posiciones {d['n_pos']:>4}")
        print(f"     coste   {fmt(d['coste']):>16} €")
        print(f"     valor   {fmt(d['valor']):>16} €")
        print(f"     latente {fmt(d['latente']):>16} €")
        if d["falta_coste"]:
            print(f"     ⚠️  SIN PRECIO: {fmt(d['falta_coste'])} € de coste excluido "
                  f"({len(d['falta_tckrs'])} posiciones)")
            print(f"        {', '.join(sorted(set(d['falta_tckrs']))[:20])}")

    # ── VALIDACION ────────────────────────────────────────────────────
    barra("VALIDACION CONTRA EL EXCEL")
    ult = serie_out[-1]
    print(f"  Fecha calculada          : {ult['fecha']}")
    print(f"  Valor cartera calculado  : {fmt(ult['valor']):>16} €")
    print(f"  Coste calculado          : {fmt(ult['coste']):>16} €")
    print(f"  LATENTE calculada        : {fmt(ult['latente']):>16} €")
    print(f"  LATENTE segun Excel (AB) : {fmt(latente_excel):>16} €")
    dif = ult["latente"] - latente_excel
    valor_excel = coste_excel + latente_excel
    dif_val = ult["valor"] - valor_excel
    dif_coste = ult["coste"] - coste_excel
    print(f"  Valor cartera Excel      : {fmt(valor_excel):>16} €")
    print()
    print(f"  Δ COSTE  (prueba dura)   : {fmt(dif_coste):>16} €")
    print(f"  Δ VALOR                  : {fmt(dif_val):>16} €   "
          f"({abs(dif_val)/valor_excel*100 if valor_excel else 0:.2f}%)")
    print(f"  Δ LATENTE                : {fmt(dif):>16} €")
    if sin_precio_hoy:
        print(f"  ⚠️  sin precio hoy: {', '.join(sorted(set(sin_precio_hoy)))}")
    print()
    # El COSTE es la prueba que importa: no depende de precios ni de la hora a la
    # que se ejecute. Si cuadra, la reconstruccion de posiciones es exacta.
    # El VALOR siempre baila un poco por el desfase entre el refresco del Excel y
    # el cierre de Yahoo, sobre todo con Wall Street abierta.
    ok_coste = abs(dif_coste) < 1.0
    ok_valor = abs(dif_val) / valor_excel < 0.01 if valor_excel else False
    print("  COSTE  :", "✅ exacto" if ok_coste else "❌ NO CUADRA -> posiciones mal reconstruidas")
    print("  VALOR  :", "✅ dentro del 1%" if ok_valor else "⚠️  fuera del 1%, revisa simbolos")
    if ok_coste and ok_valor:
        print("\n  ✅ METODO VALIDADO.")
    elif not ok_coste:
        print("\n  ❌ PARA AQUI. El coste no cuadra, algo falla en la lectura del Excel.")
    else:
        print("\n  ⚠️  El coste cuadra pero el valor se va. Mira los simbolos raros.")

    # ── PERIODOS ──────────────────────────────────────────────────────
    barra("GANANCIA POR PERIODOS (todo BRUTO)")
    lunes = hoy - timedelta(days=hoy.weekday())
    dia1 = hoy.replace(day=1)
    enero1 = date(ANIO, 1, 1)

    def latente_antes_de(ref):
        prev = [s for s in serie_out if s["fecha"] < ref]
        return prev[-1] if prev else None

    def suma(lst, desde):
        return sum(x["bruto"] for x in lst if x["fecha"] >= desde)

    for nombre, ref in (("SEMANA", lunes), ("MES", dia1), ("ANUAL", enero1)):
        base = latente_antes_de(ref)
        if base is None:
            print(f"  {nombre:<7} sin dato base anterior a {ref}")
            continue
        dl = ult["latente"] - base["latente"]
        v = suma(ventas, ref)
        d = suma(dividendos, ref)
        print(f"  {nombre:<7} base {base['fecha']} ({fmt(base['latente'])} €)")
        print(f"          Δ latentes {fmt(dl):>14} €")
        print(f"          + ventas   {fmt(v):>14} €")
        print(f"          + dividend {fmt(d):>14} €")
        print(f"          = TOTAL    {fmt(dl + v + d):>14} €\n")

    # ── GUARDAR ───────────────────────────────────────────────────────
    # JSON con la serie completa: lo consume update_from_excel_v3.py y de ahi
    # va a index.html. Guardar la serie entera (y no solo 3 numeros) hace que la
    # app pueda calcular la base de cualquier semana o mes sin recalcular nada.
    provisionales = [str(s["fecha"]) for s in serie_out if s["provisional"]]
    publicados = [s for s in serie_out if not s["incompleto"] and not s["provisional"]]
    serie_json = {
        "generado": datetime.now().isoformat(timespec="seconds"),
        # ultimo_dia es el ultimo dia PUBLICADO, o sea el ultimo cierre bueno.
        "ultimo_dia": str(publicados[-1]["fecha"]) if publicados else None,
        "ultimo_calculado": str(ult["fecha"]),
        "latente_ultimo": round(publicados[-1]["latente"], 2) if publicados else None,
        "coste_ultimo": round(publicados[-1]["coste"], 2) if publicados else None,
        "provisionales": provisionales,
        "serie": {str(s["fecha"]): round(s["latente"], 2) for s in publicados},
        "dias_descartados": [str(s["fecha"]) for s in serie_out if s["incompleto"]],
        "congelados": len(congelado),
        "salud": _salud(serie_out, dias[-1]),
    }
    SALIDA_JSON.write_text(json.dumps(serie_json, indent=1, ensure_ascii=False), encoding="utf-8")
    n_pub = len(serie_json["serie"])
    print(f"💾 Serie JSON: {n_pub} dias publicados de {len(serie_out)} calculados "
          f"en {SALIDA_JSON.name}")
    if provisionales:
        barra("⚠️  DIA NO PUBLICADO: WALL STREET SIGUE ABIERTA")
        print(f"   {', '.join(provisionales)} se ha calculado pero NO se publica.")
        print("   Los precios USA de Yahoo son ahora intradia, no cierres: si se")
        print("   guardaran, manana ESTA SEMANA restaria contra una base falsa.")
        print("   Vuelve a ejecutar despues de las 22:10 para grabar el cierre bueno.")
        print(f"   La base sigue siendo {serie_json['ultimo_dia']}, que es correcto.")

    with open(SALIDA_CSV, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh, delimiter=";")
        w.writerow(["fecha", "valor_total", "coste_total", "latente"])
        for s in serie_out:
            w.writerow([s["fecha"], f"{s['valor']:.2f}", f"{s['coste']:.2f}", f"{s['latente']:.2f}"])
    print(f"💾 Serie de {len(serie_out)} dias guardada en {SALIDA_CSV.name}")
    print("\n  Copia y pegame el bloque de VALIDACION y el de PERIODOS.")


if __name__ == "__main__":
    main()
