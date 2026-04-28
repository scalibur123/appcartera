#!/usr/bin/env python3
import openpyxl
import json
import os
import re

EXCEL_DIR = os.path.expanduser("~/AppCartera_Data")
EXCEL_FILE = os.path.join(EXCEL_DIR, "PLUSVALIAS BOLSA 26_APP.xlsm")
INDEX_FILE = os.path.expanduser("~/APPCARTERA_NUEVA/index.html")

def read_excel_to_array():
    """Lee el Excel y genera el array C"""
    if not os.path.exists(EXCEL_FILE):
        print(f"❌ Excel no encontrado: {EXCEL_FILE}")
        return None
    
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        ws = wb['2026']
    except Exception as e:
        print(f"❌ Error abriendo Excel: {e}")
        return None
    
    array_c = []
    
    # Leer desde fila 5 (después de headers en fila 4)
    for row_idx in range(5, ws.max_row + 1):
        try:
            ticker = ws[f'D{row_idx}'].value  # Columna D
            nombre = ws[f'B{row_idx}'].value  # Columna B
            moneda = ws[f'G{row_idx}'].value  # Columna G
            cambio = ws[f'J{row_idx}'].value  # Columna J (tipo de cambio)
            titulos = ws[f'K{row_idx}'].value  # Columna K
            coste_eur = ws[f'N{row_idx}'].value  # Columna N (ya en EUR)
            
            # Validar datos básicos
            if not ticker or ticker == "#VALUE!":
                continue
            
            ticker = str(ticker).strip()
            if not nombre or nombre == "#VALUE!":
                nombre = ticker
            else:
                nombre = str(nombre).strip()
            
            # Convertir a números
            try:
                titulos = float(titulos) if titulos else 0
                coste_eur = float(coste_eur) if coste_eur else 0
            except:
                continue
            
            # Ignorar si no hay títulos o es fórmula
            if titulos == 0 or isinstance(titulos, str):
                continue
            
            # Calcular precio medio en EUR
            if coste_eur > 0 and titulos > 0:
                precio_medio = coste_eur / titulos
            else:
                precio_medio = 0
            
            # SIEMPRE en EUR
            moneda_final = "EUR"
            
            item = {
                "tckr": ticker,
                "nombre": nombre,
                "titulos": round(titulos, 2),
                "coste_eur": round(coste_eur, 2),
                "precio_medio": round(precio_medio, 4),
                "moneda": moneda_final,
                "banco": "-",
                "objetivo": None
            }
            
            array_c.append(item)
            print(f"✅ {ticker}: {titulos} títulos @ €{precio_medio:.2f}")
        
        except Exception as e:
            continue
    
    return array_c

def update_index_html(array_c):
    """Actualiza el array C en index.html"""
    with open(INDEX_FILE, 'r', encoding='utf-8') as f:
        content = f.read()
    
    # Generar el nuevo array C
    array_str = json.dumps(array_c, separators=(',', ':'))
    new_const = f"const C={array_str};"
    
    # Reemplazar el array C
    pattern = r'const C=\[.*?\];'
    new_content = re.sub(pattern, new_const, content, flags=re.DOTALL)
    
    with open(INDEX_FILE, 'w', encoding='utf-8') as f:
        f.write(new_content)
    
    print(f"\n✅ index.html actualizado con {len(array_c)} valores")

if __name__ == "__main__":
    print("📖 Leyendo Excel...")
    array_c = read_excel_to_array()
    
    if array_c and len(array_c) > 0:
        print(f"\n✅ {len(array_c)} valores encontrados")
        print("📝 Actualizando index.html...")
        update_index_html(array_c)
        print("✅ ¡HECHO!")
    else:
        print("❌ No se encontraron valores")

