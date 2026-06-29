#!/usr/bin/env python3
"""
update_from_excel_v3.py - SOLUCION DEFINITIVA

Lee el Excel y resuelve simbolos Yahoo automaticamente desde la columna B
(que tiene formato "Nombre (MIC:TICKER)" tipo "Repsol SA (BMEX:REP)").

ESTRATEGIA EN CASCADA:
  1. Si tickers_override.json tiene el ticker -> usar ese (override manual gana siempre)
  2. Leer columna B o AO buscando patron (MIC:TICKER) y mapear MIC -> sufijo Yahoo
  3. Si la columna B/AO esta vacia o #VALUE! -> usar moneda como heuristica:
     - USD -> ticker tal cual
     - EUR sin sufijo -> .MC (Madrid) por defecto
  4. Si nada funciona -> dejar ticker tal cual y avisar

NOTA IMPORTANTE: La columna B es una formula que da #VALUE! al leer con Python
si el Excel no se ha guardado con valores cacheados. SOLUCION: copiar la columna B
como VALOR a la columna AO (script export_mic_values.py lo hace automaticamente).
"""

import openpyxl
import json
import re
import sys
from pathlib import Path
from datetime import datetime

HOME = Path.home()
PROYECTO = HOME / "APPCARTERA_NUEVA"
EXCEL = Path.home() / "Library/Mobile Documents/com~apple~CloudDocs/INVERSION/PLUSVALIAS BOLSA 26.xlsm"
HOJA = "2026"
FILA_INI = 5
FILA_FIN = 258
INDEX_HTML = PROYECTO / "index.html"
TICKERS_JSON = PROYECTO / "tickers.json"
OVERRIDE_JSON = PROYECTO / "tickers_override.json"
MIC_NAMES_TXT = PROYECTO / "mic_nombres.txt"  # opcional: pegar ahi columna B manualmente

# === MAPA MIC -> SUFIJO YAHOO ===
MIC_TO_YAHOO_SFX = {
    "BMEX": ".MC",   # Madrid (Bolsas y Mercados Españoles)
    "XMAD": ".MC",   # Madrid (codigo alternativo)
    "XPAR": ".PA",   # Paris (Euronext Paris)
    "XAMS": ".AS",   # Amsterdam (Euronext Amsterdam)
    "XBRU": ".BR",   # Bruselas (Euronext Brussels)
    "XLIS": ".LS",   # Lisboa (Euronext Lisbon)
    "XHEL": ".HE",   # Helsinki (Nasdaq Helsinki)
    "XSTO": ".ST",   # Estocolmo
    "XCSE": ".CO",   # Copenhague
    "XOSL": ".OL",   # Oslo
    "XFRA": ".DE",   # Frankfurt
    "XETR": ".DE",   # Xetra (Frankfurt electronico)
    "XMIL": ".MI",   # Milan
    "XLON": ".L",    # Londres
    "XWBO": ".VI",   # Viena
    "XSWX": ".SW",   # Suiza (SIX)
    "XVTX": ".VX",   # Suiza Virt-x
    "XWAR": ".WA",   # Varsovia
    "XPRA": ".PR",   # Praga
    "XBUD": ".BD",   # Budapest
    "XATH": ".AT",   # Atenas
    "XIST": ".IS",   # Estambul
    "XTAE": ".TA",   # Tel Aviv
    # USA: sin sufijo
    "XNAS": "",      # Nasdaq
    "XNYS": "",      # NYSE
    "XASE": "",      # NYSE American (AMEX)
    "BATS": "",      # BATS / Cboe
    "ARCX": "",      # NYSE Arca
    "XOTC": "",      # OTC
    "PINX": "",      # Pink Sheets
    # Otros
    "XTSX": ".V",    # TSX Venture
    "XTSE": ".TO",   # Toronto
    "XHKG": ".HK",   # Hong Kong
    "XTKS": ".T",    # Tokio
    "XSHG": ".SS",   # Shanghai
    "XSHE": ".SZ",   # Shenzhen
}

REGEX_MIC = re.compile(r"\(([A-Z][A-Z0-9]{2,4}):([A-Z0-9]+)\)")


BANCO_NORMALIZE = {
    'r4': 'R4', 'R4': 'R4',
    'ing': 'ING', 'Ing': 'ING', 'ING': 'ING',
    'ibkr': 'IBKR', 'IBKR': 'IBKR', 'Ibkr': 'IBKR',
    'revolut': 'REVOLUT', 'Revolut': 'REVOLUT',
    'medio': 'MEDIOLANUM', 'MEDIO': 'MEDIOLANUM', 'Medio': 'MEDIOLANUM',
    'myinv': 'MYINV', 'Myinv': 'MYINV', 'MyInv': 'MYINV',
    'r4/ing': 'R4/ING', 'R4/ING': 'R4/ING', 'ing/r4': 'R4/ING',
    'medio/ing': 'MEDIO/ING', 'MEDIO/ING': 'MEDIO/ING', 'ing/medio': 'MEDIO/ING',
}

def normalizar_banco(b):
    if not b: return '-'
    b = str(b).strip()
    return BANCO_NORMALIZE.get(b, b.upper())


def cargar_overrides():
    if OVERRIDE_JSON.exists():
        with open(OVERRIDE_JSON, "r", encoding="utf-8") as f:
            data = json.load(f)
        return {k: v for k, v in data.items() if not k.startswith("_")}
    return {}


def cargar_mic_externos():
    """
    Si existe mic_nombres.txt en el proyecto, lo lee como respaldo.
    Es un fichero con una linea por ticker: "Nombre (MIC:TICKER)".
    El orden NO importa. Construye un dict ticker -> mic.
    """
    if not MIC_NAMES_TXT.exists():
        return {}
    mapa = {}
    with open(MIC_NAMES_TXT, "r", encoding="utf-8") as f:
        for linea in f:
            m = REGEX_MIC.search(linea)
            if m:
                mic, tckr = m.group(1), m.group(2)
                mapa[tckr] = mic
    return mapa


def leer_diana(wb, tickers_vivos):
    if 'DIANA' not in wb.sheetnames:
        return {}
    ws = wb['DIANA']
    objetivos = {}
    for row in range(12, 221):
        ticker = ws[f'B{row}'].value
        objetivo = ws[f'AC{row}'].value
        if not ticker or not isinstance(ticker, str): continue
        ticker = ticker.strip()
        if ticker not in tickers_vivos: continue
        if ticker in objetivos: continue
        if not isinstance(objetivo, (int, float)): continue
        objetivos[ticker] = round(float(objetivo), 4)
    return objetivos


def leer_proximas_compras(wb):
    """
    Lee las proximas compras de la pestana 'Compras'.
    Fila 1 = cabecera. Filas 2+ = valores hasta encontrar fila vacia en col B.
    Columnas (segun cabecera fila 1):
      B  = simbolo Yahoo directamente (ej: VCT.PA, MTX.DE, HOOD)
      R  = banco
      U  = precio condicion (ej: ">67,9")
      V  = precio maximo de compra
      Y  = objetivo personal
    El ticker se extrae del simbolo: quitar sufijo (.PA, .DE, etc.)
    """
    hoja = 'Compras'
    if hoja not in wb.sheetnames:
        print(f'⚠️  Pestaña "{hoja}" no encontrada en el Excel')
        return []
    ws = wb[hoja]
    proximas = []

    _overrides = cargar_overrides()

    for row in range(2, 500):  # fila 2 en adelante, hasta vacío
        symbol_raw = ws.cell(row=row, column=2).value   # B = ticker
        if not symbol_raw or not isinstance(symbol_raw, str) or not symbol_raw.strip():
            break  # primera fila vacía = fin de lista

        symbol = symbol_raw.strip()

        # Extraer ticker limpio (sin sufijo de mercado)
        tckr = symbol.split('.')[0]

        # Resolver símbolo Yahoo en cascada:
        # 1. Si ya tiene sufijo (.DE, .PA, etc.) -> usar tal cual
        # 2. Override manual en tickers_override.json
        # 3. Columna A: "Nombre (MIC:TICKER)" -> sufijo Yahoo via MIC_TO_YAHOO_SFX
        # 4. Búsqueda automática en Yahoo Finance (último recurso)
        if '.' not in symbol and '=' not in symbol:
            if tckr in _overrides:
                symbol = _overrides[tckr]
            else:
                nombre_raw = ws.cell(row=row, column=1).value  # A = nombre con MIC
                mic_encontrado = False
                if nombre_raw and isinstance(nombre_raw, str):
                    m = REGEX_MIC.search(nombre_raw)
                    if m:
                        mic = m.group(1)
                        if mic in MIC_TO_YAHOO_SFX:
                            symbol = tckr + MIC_TO_YAHOO_SFX[mic]
                            guardar_override(tckr, symbol)
                            print(f"   ✅ {tckr} -> {symbol} (MIC:{mic} desde col A)")
                            mic_encontrado = True
                if not mic_encontrado:
                    print(f"   🔍 {tckr} sin MIC en col A, buscando en Yahoo Finance...")
                    sym_auto = buscar_simbolo_yahoo_auto(tckr)
                    if sym_auto:
                        symbol = sym_auto
                        guardar_override(tckr, sym_auto)
                        print(f"   ✅ {tckr} -> {sym_auto} (Yahoo, guardado en overrides)")
                    else:
                        print(f"   ⚠️  {tckr} no resuelto — quedará sin precio")

        banco_raw    = ws.cell(row=row, column=21).value  # U = Banco
        p_cond_raw   = ws.cell(row=row, column=24).value  # X = P. condición
        p_max_raw    = ws.cell(row=row, column=25).value  # Y = Precio Max.
        objetivo_raw = ws.cell(row=row, column=8).value   # H = Investing (objetivo)

        def fmt_val(v):
            if v is None or v == '-' or (isinstance(v, str) and ('#' in v or not v.strip())):
                return None
            if isinstance(v, (int, float)):
                return round(float(v), 4)
            return str(v).strip()

        banco   = str(banco_raw).strip() if banco_raw and str(banco_raw).strip() not in ('-', '') else None
        p_cond  = fmt_val(p_cond_raw)
        p_max   = fmt_val(p_max_raw)
        objetivo = fmt_val(objetivo_raw)

        proximas.append({
            'tckr':    tckr,
            'symbol':  symbol,
            'banco':   banco,
            'p_cond':  p_cond,
            'p_max':   p_max,
            'objetivo': objetivo,
        })

    print(f'✅ {len(proximas)} proximas compras leidas de pestana Compras')
    for p in proximas:
        print(f"   {p['tckr']} ({p['symbol']}) banco={p['banco']} p_max={p['p_max']} cond={p['p_cond']}")
    return proximas


def leer_excel_con_mic():
    """
    Lee Excel agrupando por ticker. Intenta extraer MIC de:
      - columna B (formato "Nombre (MIC:TICKER)")
      - columna AO (si la macro VBA copio los valores ahi)
      - mic_nombres.txt (fallback manual)
    """
    if not EXCEL.exists():
        print(f"❌ Excel no encontrado: {EXCEL}")
        sys.exit(1)

    # Cargar dos versiones: una con valores cacheados y otra con formulas
    wb = openpyxl.load_workbook(EXCEL, data_only=True, keep_vba=False)
    ws = wb[HOJA]
    mensual_data = None
    plusv_hoy = 0
    plusv_semana = 0
    # Leer datos pestaña Mensual
    try:
        ws_m = wb['Mensual']
        bruto_anual = ws_m['R13'].value or 0
        neto_anual = ws_m['S13'].value or 0
        equiv_bruto = ws_m['S15'].value or 0
        neto_mensual = ws_m['J13'].value or 0
        # Plusvalías realizadas esta semana (no incluidas en F57 hasta el sábado)
        from datetime import date, timedelta
        hoy_d = date.today()
        lunes_d = hoy_d - timedelta(days=hoy_d.weekday())
        plusv_semana = 0
        ws_ej = wb['2026']
        for row in range(262, 401):
            fecha_v = ws_ej.cell(row=row, column=17).value
            plusv = ws_ej.cell(row=row, column=25).value
            if fecha_v and hasattr(fecha_v, 'date') and fecha_v.date() >= lunes_d and plusv:
                plusv_semana += plusv
        sueldo_mensual_bruto = ws_m['N13'].value or 0
        neto_anual_nomina = neto_mensual * 14

        # Promedios netos 2024/2026 desde pestaña Histórico
        ws_hist = wb['Histórico']
        prom_neto_conservador = ws_hist['F30'].value or 0  # 14 pagas
        prom_neto_optimista   = ws_hist['J30'].value or 0  # 12 pagas

        def calcular_bruto_desde_neto(neto_anual):
            ss = 0.034
            tramos_e = [(12450,0.095),(7750,0.12),(15000,0.15),(24800,0.185),(120000,0.225),(9e9,0.235)]
            tramos_a = [(12450,0.105),(7750,0.12),(15000,0.14),(24800,0.175),(120000,0.2125),(9e9,0.2375)]
            min_ded = 5550 + 1200 + 700
            def neto_de_bruto(b):
                base = b*(1-ss) - min_ded
                ce = ca = 0; re_ = ra = base
                for lim,t in tramos_e:
                    if re_<=0: break
                    x=min(re_,lim); ce+=x*t; re_-=x
                for lim,t in tramos_a:
                    if ra<=0: break
                    x=min(ra,lim); ca+=x*t; ra-=x
                return b*(1-ss)-ce-ca
            lo,hi = neto_anual*0.8,neto_anual*2
            for _ in range(60):
                mid=(lo+hi)/2
                if neto_de_bruto(mid)<neto_anual: lo=mid
                else: hi=mid
            return round((lo+hi)/2)

        equiv_bruto_calculado = calcular_bruto_desde_neto(neto_anual_nomina)
        
        # Leer compras desglosadas (col D=ticker, I=fecha, K=titulos, N=coste_eur)
        compras_por_ticker = {}
        for row in range(5, 258):
            tckr_r = ws.cell(row=row, column=4).value
            fecha_i = ws.cell(row=row, column=9).value
            titulos_i = ws.cell(row=row, column=11).value
            coste_i = ws.cell(row=row, column=14).value
            if tckr_r and fecha_i and titulos_i and coste_i:
                tk = str(tckr_r).strip()
                if tk not in compras_por_ticker:
                    compras_por_ticker[tk] = []
                fecha_str = fecha_i.strftime('%d/%m/%Y') if hasattr(fecha_i, 'strftime') else str(fecha_i)[:10]
                precio_compra = round(coste_i / titulos_i, 4) if titulos_i else 0
                compras_por_ticker[tk].append({
                    'fecha': fecha_str,
                    'titulos': titulos_i,
                    'precio': precio_compra,
                    'coste': round(coste_i, 2)
                })

        def fmt_eur(v):
            return f"{v:,.2f} €".replace(",", "X").replace(".", ",").replace("X", ".")
        
        mensual_data = {
            'bruto_anual': fmt_eur(bruto_anual),
            'neto_anual': fmt_eur(neto_anual),
            'equiv_bruto': fmt_eur(equiv_bruto_calculado),
            'neto_nomina': fmt_eur(neto_anual_nomina),
            'sueldo_bruto': fmt_eur(sueldo_mensual_bruto),
            'sueldo_neto': fmt_eur(neto_mensual),
            'prom_neto_conservador': fmt_eur(prom_neto_conservador),
            'prom_neto_optimista': fmt_eur(prom_neto_optimista),
            'plusv_semana': round(plusv_semana, 2),
            'plusv_hoy': round(plusv_hoy, 2),
        }
    except Exception as e:
        print(f"Warning: no se pudo leer pestaña Mensual: {e}")
        mensual_data = None


    mic_externo = cargar_mic_externos()
    if mic_externo:
        print(f"📄 Cargados {len(mic_externo)} mapeos MIC desde {MIC_NAMES_TXT.name}")

    posiciones = {}
    orden = []
    sin_mic = []

    for row in range(FILA_INI, FILA_FIN + 1):
        ticker = ws[f"D{row}"].value
        if not ticker or not str(ticker).strip():
            continue
        ticker = str(ticker).strip()
        titulos = ws[f"K{row}"].value or 0
        coste_eur = ws[f"N{row}"].value or 0
        moneda = ws[f"G{row}"].value
        precio_excel = ws[f"E{row}"].value

        # Intentar extraer MIC: B -> AO -> mic_nombres.txt
        mic = None
        nombre_completo = None

        for col_letter in ["B", "AO"]:
            v = ws[f"{col_letter}{row}"].value
            if v and isinstance(v, str) and "#" not in v:
                m = REGEX_MIC.search(v)
                if m:
                    mic = m.group(1)
                    nombre_completo = v.strip()
                    break

        if not mic and ticker in mic_externo:
            mic = mic_externo[ticker]

        if ticker not in posiciones:
            posiciones[ticker] = {
                "tckr": ticker,
                "nombre": nombre_completo or ticker,
                "titulos": 0,
                "coste_eur": 0,
                "moneda": moneda,
                "precio_excel": precio_excel,
                "mic": mic,
                "banco": normalizar_banco(ws[f"H{row}"].value),
                "objetivo": None,
            }
            orden.append(ticker)
            if not mic:
                sin_mic.append(ticker)

        posiciones[ticker]["titulos"] += titulos
        posiciones[ticker]["coste_eur"] += coste_eur
        # Si en una fila posterior aparece el MIC, actualizar
        if mic and not posiciones[ticker]["mic"]:
            posiciones[ticker]["mic"] = mic
            posiciones[ticker]["nombre"] = nombre_completo or ticker
            if ticker in sin_mic:
                sin_mic.remove(ticker)

    # Leer objetivos de pestana DIANA
    tickers_vivos = set(posiciones.keys())
    objetivos = leer_diana(wb, tickers_vivos)
    for t, obj in objetivos.items():
        if t in posiciones:
            posiciones[t]['objetivo'] = obj
    print(f'OK {len(objetivos)} objetivos leidos de pestana DIANA')

    # Leer proximas compras de pestana DIANA
    proximas_compras = leer_proximas_compras(wb)

    # DEBUG
    fubo = posiciones.get('FUBO')
    abeo = posiciones.get('ABEO')
    if fubo: print(f'DEBUG FUBO en posiciones: objetivo={fubo.get("objetivo")}, banco={fubo.get("banco")}')
    if abeo: print(f'DEBUG ABEO en posiciones: objetivo={abeo.get("objetivo")}, banco={abeo.get("banco")}')

    return [posiciones[t] for t in orden], sin_mic, mensual_data, compras_por_ticker, proximas_compras


def resolver_simbolo_yahoo(p, overrides):
    """Resuelve simbolo Yahoo en cascada: override > mic > moneda > tal cual."""
    ticker = p["tckr"]

    # 1. Override manual
    if ticker in overrides:
        return overrides[ticker], "override"

    # 2. Si el ticker ya tiene sufijo punto, usar tal cual
    if "." in ticker or "=" in ticker:
        return ticker, "ya_tenia_sufijo"

    # 3. MIC desde Excel
    mic = p.get("mic")
    if mic and mic in MIC_TO_YAHOO_SFX:
        sfx = MIC_TO_YAHOO_SFX[mic]
        return ticker + sfx, f"mic_{mic}"

    # 4. MIC desconocido -> avisar pero seguir
    if mic:
        return ticker, f"mic_desconocido_{mic}"

    # 5. Heuristica por moneda
    if p.get("moneda") == "USD":
        return ticker, "fallback_usd_directo"
    if p.get("moneda") == "EUR":
        return ticker + ".MC", "fallback_madrid"

    return ticker, "sin_resolver"


def buscar_simbolo_yahoo_auto(ticker):
    """
    Busca el simbolo correcto en Yahoo Finance dado un ticker.
    Usa el endpoint de busqueda de Yahoo. Devuelve el simbolo o None si no encuentra.
    Guarda el resultado en tickers_override.json automaticamente para no volver a buscar.
    """
    import urllib.request, urllib.parse, json as _json, time
    url = f"https://query1.finance.yahoo.com/v1/finance/search?q={urllib.parse.quote(ticker)}&quotesCount=5&newsCount=0"
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=8) as r:
            data = _json.loads(r.read().decode())
        quotes = data.get("finance", {}).get("result", [{}])[0].get("quotes", [])
        if not quotes:
            return None
        # Buscar coincidencia exacta de ticker primero
        for q in quotes:
            sym = q.get("symbol", "")
            if sym.upper() == ticker.upper():
                return sym
        # Si no hay exacta, coger el primero que sea equity
        for q in quotes:
            if q.get("quoteType") in ("EQUITY", "ETF"):
                return q.get("symbol")
        return quotes[0].get("symbol")
    except Exception as e:
        print(f"   ⚠️  Error buscando {ticker} en Yahoo: {e}")
        return None


def guardar_override(ticker, symbol):
    """Guarda un override en tickers_override.json."""
    import json as _json
    data = {}
    if OVERRIDE_JSON.exists():
        with open(OVERRIDE_JSON, "r", encoding="utf-8") as f:
            data = _json.load(f)
    data[ticker] = symbol
    with open(OVERRIDE_JSON, "w", encoding="utf-8") as f:
        _json.dump(data, f, indent=2, ensure_ascii=False)


def construir_const_C_compacta(posiciones, ticker_map, compras_por_ticker={}):
    items = []
    for p in posiciones:
        tckr = p["tckr"]
        sym = ticker_map[tckr]["yahoo"]
        coste = round(p["coste_eur"], 4)
        titulos = round(p["titulos"], 4) if p["titulos"] != int(p["titulos"]) else int(p["titulos"])
        precio_medio = round(coste / titulos, 4) if titulos else 0
        nombre = p.get("nombre", tckr)
        if "(" in nombre:
            nombre = nombre.split("(")[0].strip()
        if not nombre or "#" in nombre:
            nombre = tckr
        item_json = {
            "tckr": tckr,
            "nombre": nombre,
            "titulos": float(titulos),
            "coste_eur": coste,
            "precio_medio": precio_medio,
            "moneda": p["moneda"],
            "banco": p.get("banco", "-"),
            "objetivo": p.get("objetivo"),
            "symbol": sym,
            "compras": compras_por_ticker.get(tckr, []),
        }
        items.append(json.dumps(item_json, ensure_ascii=False, separators=(",", ":")))
    return "const C=[" + ",".join(items) + "];"


def asegurar_proximas_compras(html, proximas_compras):
    """
    Inyecta o reemplaza la pestana 'proximas' en index.html con los datos
    de proximas compras leidos del Excel (DIANA filas 212-215).
    Si la pestana no existe la crea. Si ya existe la reemplaza.
    """
    import json as _json
    proximas_json = _json.dumps(proximas_compras, ensure_ascii=False, separators=(',', ':'))

    # ---- Boton de navegacion ----
    btn = '<button class="nav-btn" onclick="showTab(\'proximas\',this)">Próximas</button>'

    # ---- Seccion HTML de la pestana ----
    section = '''<div id="proximas" class="section">
  <div style="padding:16px">
    <h2 style="margin-bottom:12px;font-size:14px;font-weight:700;color:var(--muted);letter-spacing:1px;text-transform:uppercase">Próximas compras</h2>
    <div id="proximas-list"></div>
  </div>
</div>'''

    # ---- Funcion JS de renderizado ----
    render_fn = ("""
(function(){
  var PROXIMAS=%%PROXIMAS_JSON%%;

  function fmtPrecio(v){
    if(v==null)return '-';
    return v.toLocaleString('es-ES',{minimumFractionDigits:2,maximumFractionDigits:2});
  }

  function buildCard(p, precioReal, pct){
    var condicional=p.p_cond&&p.p_cond!==null;
    var badgeCond=condicional
      ?'<span style="display:inline-block;padding:2px 7px;border-radius:10px;font-size:10px;font-weight:700;background:rgba(255,165,0,0.15);color:#f5a623;margin-left:6px">COND</span>'
      :'<span style="display:inline-block;padding:2px 7px;border-radius:10px;font-size:10px;font-weight:700;background:rgba(76,175,80,0.15);color:var(--green);margin-left:6px">DIRECTA</span>';
    var bancoHtml=p.banco
      ?'<span style="font-size:11px;color:var(--muted);margin-left:6px">'+p.banco+'</span>'
      :'';
    var condHtml=condicional
      ?'<div style="margin-top:4px;font-size:12px;color:#f5a623">⚡ Condición: <b>'+p.p_cond+'</b></div>'
      :'';
    var pmaxHtml=p.p_max!=null
      ?'<div style="margin-top:3px;font-size:12px;color:var(--muted)">Precio máx: <b style="color:var(--fg)">'+fmtPrecio(p.p_max)+'</b></div>'
      :'';
    var objHtml=p.objetivo!=null
      ?'<div style="margin-top:3px;font-size:12px;color:var(--muted)">Objetivo: <b style="color:var(--green)">'+fmtPrecio(p.objetivo)+'</b></div>'
      :'';
    var pctStr='';
    if(pct!=null){
      var pctColor=pct>=0?'var(--green)':'var(--red)';
      var pctSign=pct>=0?'+':'';
      pctStr='<span style="font-size:12px;color:'+pctColor+';margin-left:8px">'+pctSign+pct.toFixed(2)+'%</span>';
    }
    var precioHtml=precioReal!=null
      ?'<div style="margin-top:5px;font-size:14px;color:var(--muted)">Precio actual: <b style="color:var(--fg);font-size:16px">'+fmtPrecio(precioReal)+'</b>'+pctStr+'</div>'
      :'<div style="margin-top:5px;font-size:12px;color:var(--muted)">Precio actual: <i>cargando...</i></div>';
    return '<div style="padding:14px 0;border-bottom:1px solid var(--border)">'
      +'<div style="display:flex;align-items:center;flex-wrap:wrap;margin-bottom:2px">'
      +'<span style="font-size:16px;font-weight:700;color:var(--fg)">'+p.tckr+'</span>'
      +bancoHtml+badgeCond+'</div>'
      +condHtml+precioHtml+pmaxHtml+objHtml
      +'</div>';
  }

  async function renderProximas(){
    var el=document.getElementById('proximas-list');
    if(!el)return;
    if(!PROXIMAS||PROXIMAS.length===0){
      el.innerHTML='<p style="color:var(--muted);font-size:14px">Sin próximas compras registradas.</p>';
      return;
    }
    // Mostrar skeleton mientras carga
    el.innerHTML=PROXIMAS.map(function(p){return buildCard(p,null,null);}).join('');
    // Fetch precios reales al servidor (mismo endpoint que usa la cartera)
    var symbols=PROXIMAS.map(function(p){return p.symbol;}).join(',');
    try{
      var r=await fetch('/?symbols='+encodeURIComponent(symbols));
      var data=await r.json();
      var priceMap={};
      if(data&&data.quoteResponse&&data.quoteResponse.result){
        data.quoteResponse.result.forEach(function(q){
          priceMap[q.symbol]={price:q.regularMarketPrice,pct:q.regularMarketChangePercent};
        });
      }
      el.innerHTML=PROXIMAS.map(function(p){
        var q=priceMap[p.symbol];
        return buildCard(p, q?q.price:null, q?q.pct:null);
      }).join('');
    }catch(e){
      // Si falla el fetch, mostrar sin precio
      el.innerHTML=PROXIMAS.map(function(p){return buildCard(p,null,null);}).join('');
    }
  }

  var _origShowTab=window.showTab;
  window.showTab=function(id,btn){
    _origShowTab&&_origShowTab(id,btn);
    if(id==='proximas')renderProximas();
  };
})();
""").replace('%%PROXIMAS_JSON%%', proximas_json)

    # Insertar boton de nav si no existe (antes del boton de Analistas o al final de nav)
    if "onclick=\"showTab('proximas'" not in html:
        if "<button class=\"nav-btn\" onclick=\"showTab('analistas'" in html:
            html = html.replace(
                "<button class=\"nav-btn\" onclick=\"showTab('analistas'",
                btn + '\n  <button class="nav-btn" onclick="showTab(\'analistas\'',
                1
            )
        else:
            html = html.replace('</nav>', '  ' + btn + '\n</nav>', 1)

    # Insertar seccion HTML si no existe (antes de analistas o antes de </body>)
    if 'id="proximas"' not in html:
        if '<div id="analistas"' in html:
            html = html.replace('<div id="analistas"', section + '\n<div id="analistas"', 1)
        else:
            html = html.replace('</body>', section + '\n</body>', 1)

    # Inyectar/reemplazar el bloque JS (marcado con comentario para poder reemplazarlo)
    marker_start = '<!--proximas-compras-start-->'
    marker_end = '<!--proximas-compras-end-->'
    js_block = marker_start + '<script>' + render_fn + '</script>' + marker_end

    if marker_start in html:
        import re as _re
        html = _re.sub(
            re.escape(marker_start) + r'.*?' + re.escape(marker_end),
            js_block,
            html,
            flags=re.DOTALL
        )
    else:
        html = html.replace('</body>', js_block + '\n</body>', 1)

    # Actualizar array de tabs para swipe si no tiene 'proximas'
    import re as _re2
    def add_proximas_to_tabs(m):
        tabs_str = m.group(0)
        if 'proximas' in tabs_str:
            return tabs_str
        if "'analistas'" in tabs_str:
            return tabs_str.replace("'analistas'", "'proximas','analistas'")
        return tabs_str.replace(']', ",'proximas']")
    html = _re2.sub(r"const tabs=\[[^\]]+\]", add_proximas_to_tabs, html)

    return html


def asegurar_historico(html):
    """Añade pestaña Historico y funcion loadHistorico si no están"""
    if 'loadHistorico' not in html:
        html = html.replace(
            "<button class=\"nav-btn\" onclick=\"showTab('mensual',this)\">Mensual</button>",
            "<button class=\"nav-btn\" onclick=\"showTab('mensual',this)\">Mensual</button>\n  <button class=\"nav-btn\" onclick=\"showTab('historico',this);loadHistorico()\">Historico</button>"
        )
        historico_section = '''<div id="historico" class="section">
  <div style="padding:16px">
    <h2 style="margin-bottom:12px;font-size:16px;color:var(--muted)">Historico de alertas</h2>
    <div id="hist-list">Cargando...</div>
  </div>
</div>'''
        load_fn = """
window.loadHistorico = async function() {
  const el = document.getElementById('hist-list');
  try {
    const r = await fetch('/historico');
    const data = await r.json();
    if (data.length === 0) { el.innerHTML = '<p style="color:var(--muted)">Sin registros aun.</p>'; return; }
    el.innerHTML = data.map(d => {
      const fecha = new Date(d.fecha).toLocaleString('es-ES');
      const color = d.evento.includes('salio') ? 'var(--red)' : 'var(--green)';
      return '<div style="padding:10px;border-bottom:1px solid var(--border)"><span style="color:'+color+';font-weight:600">' + d.ticker + '</span> <span style="color:var(--muted);font-size:12px">' + d.banco + '</span><br><span style="font-size:13px">' + d.evento + ' - ' + d.precio + '</span><br><span style="color:var(--muted);font-size:11px">' + fecha + '</span></div>';
    }).join('');
  } catch(e) { el.innerHTML = 'Error cargando historico'; }
};
"""
        html = html.replace('</body>', historico_section + '\n</body>')
        html = html.replace('</script>\n</body>', load_fn + '</script>\n</body>')
    return html


def leer_rendimiento_cartera():
    import openpyxl
    from datetime import datetime
    wb = openpyxl.load_workbook(open(str(EXCEL), 'rb'), read_only=True, data_only=True)
    ws = wb['Semanal']
    meses = {'Enero':1,'Febrero':2,'Marzo':3,'Abril':4,'Mayo':5,'Junio':6,
             'Julio':7,'Agosto':8,'Septiembre':9,'Octubre':10,'Noviembre':11,'Diciembre':12}
    mes_hoy = datetime.now().month
    total_mes = 0
    total_anual = 0
    for row in ws.iter_rows(min_row=4, max_row=60, values_only=True):
        if row[3] is not None and isinstance(row[5], (int,float)) and row[5] != 0:
            if meses.get(row[3], 0) == mes_hoy:
                total_mes = row[5]
            total_anual += row[5]
    ws_irpf = wb['IRPF']
    tipo_irpf = ws_irpf.cell(35, 6).value or 0.20925
    return {'mes': total_mes, 'mes_n': total_mes*(1-tipo_irpf), 'anual': total_anual, 'anual_n': total_anual*(1-tipo_irpf)}

def leer_ganancias_realizadas():
    import openpyxl
    from datetime import datetime, timedelta
    wb = openpyxl.load_workbook(open(str(EXCEL), "rb"), read_only=True, data_only=True)
    hoy = datetime.now()
    inicio_semana = (hoy - timedelta(days=hoy.weekday())).replace(hour=0,minute=0,second=0,microsecond=0)
    inicio_mes = hoy.replace(day=1,hour=0,minute=0,second=0,microsecond=0)
    r = {"sem_b":0,"sem_n":0,"mes_b":0,"mes_n":0,"ani_b":0,"ani_n":0}

    ws = wb["2026"]
    for row in ws.iter_rows(min_row=262, max_row=400, values_only=True):
        fecha, bruta, neta = row[16], row[24], row[25]
        if not fecha or not isinstance(fecha, datetime) or not bruta or not neta: continue
        if fecha.year != hoy.year: continue
        r["ani_b"] += bruta; r["ani_n"] += neta
        if fecha >= inicio_mes: r["mes_b"] += bruta; r["mes_n"] += neta
        if fecha >= inicio_semana: r["sem_b"] += bruta; r["sem_n"] += neta

    ws2 = wb["dividendos 26"]
    for row in ws2.iter_rows(min_row=5, max_row=500, values_only=True):
        fecha, bruta, neta = row[0], row[7], row[8]
        if not fecha or not isinstance(fecha, datetime) or not bruta or not neta: continue
        if fecha.year != hoy.year: continue
        r["ani_b"] += bruta; r["ani_n"] += neta
        if fecha >= inicio_mes: r["mes_b"] += bruta; r["mes_n"] += neta
        if fecha >= inicio_semana: r["sem_b"] += bruta; r["sem_n"] += neta

    return r

def actualizar_index_html(const_C_linea, mensual_data=None, ganancias_data=None, rendimiento_data=None, proximas_compras=None, ventas_data=None):
    if not INDEX_HTML.exists():
        print(f"❌ index.html no encontrado: {INDEX_HTML}")
        sys.exit(1)
    html = INDEX_HTML.read_text(encoding="utf-8")
    pattern = re.compile(r"const\s+C\s*=\s*\[.*?\]\s*;", re.DOTALL)
    if not pattern.search(html):
        print("❌ No se encontro 'const C=[...]' en index.html")
        sys.exit(1)
    nuevo_html = pattern.sub(const_C_linea, html, count=1)
    fecha_hoy = datetime.now().strftime("%-d %b %Y").replace("Jan","ene").replace("Feb","feb").replace("Mar","mar").replace("Apr","abr").replace("May","may").replace("Jun","jun").replace("Jul","jul").replace("Aug","ago").replace("Sep","sep").replace("Oct","oct").replace("Nov","nov").replace("Dec","dic")
    nuevo_html = re.sub(r"Datos del Excel del [^.]+\.", f"Datos del Excel del {fecha_hoy}.", nuevo_html)
    version = datetime.now().strftime("%Y%m%d%H%M%S")
    nuevo_html = re.sub(r'content="[0-9.]+"(?=[^>]*name="app-version"|(?<=app-version")[^>]*)>', f'content="{version}">', nuevo_html)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = INDEX_HTML.parent / f"index.html.backup_{timestamp}"
    backup.write_text(html, encoding="utf-8")

    # Actualizar HTML pestaña Mensual
    if mensual_data:
        import re as re2
        nuevo_html = re2.sub(
            r'(Bruto</span><span class="mensual-val green">)[^<]*(</span>)',
            f'\g<1>{mensual_data["bruto_anual"]}\g<2>',
            nuevo_html
        )
        nuevo_html = re2.sub(
            r'(<span class="plusv-label">Neto</span><span class="mensual-val green">)[^<]*(</span>)',
            f'\g<1>{mensual_data["neto_anual"]}\g<2>',
            nuevo_html
        )
        nuevo_html = re2.sub(
            r'(Equivale a</span><span class="mensual-val green">)[^<]*(</span>)',
            f'\g<1>{mensual_data["equiv_bruto"]} brutos\g<2>',
            nuevo_html
        )
        nuevo_html = re2.sub(
            r'(id="sueldo-bruto-val">)[^<]*(</span>)',
            f'\g<1>{mensual_data["sueldo_bruto"]}\g<2>',
            nuevo_html
        )
        nuevo_html = re2.sub(
            r'(id="sueldo-neto-val">)[^<]*(</span>)',
            f'\g<1>{mensual_data["sueldo_neto"]}\g<2>',
            nuevo_html
        )
        # Inyectar promedios netos 14 pagas 24/26
        nuevo_html = re2.sub(
            r'(id="prom-neto-conservador-val">)[^<]*(</span>)',
            f'\g<1>{mensual_data["prom_neto_conservador"]}\g<2>',
            nuevo_html
        )
        nuevo_html = re2.sub(
            r'(id="prom-neto-optimista-val">)[^<]*(</span>)',
            f'\g<1>{mensual_data["prom_neto_optimista"]}\g<2>',
            nuevo_html
        )
        # Si los elementos no existen aún en el HTML, añadirlos tras sueldo-neto
        if 'prom-neto-conservador-val' not in nuevo_html:
            bloque_prom = (
                '<div class="plusv-row"><span class="plusv-label">Prom. neto 14p 24/26</span>'
                '<span class="mensual-val green" id="prom-neto-conservador-val">' + mensual_data['prom_neto_conservador'] + '</span></div>'
                '<div class="plusv-row"><span class="plusv-label">Prom. neto 12p 24/26</span>'
                '<span class="mensual-val green" id="prom-neto-optimista-val">' + mensual_data['prom_neto_optimista'] + '</span></div>'
            )
            nuevo_html = re2.sub(
                r'(id="sueldo-neto-val">[^<]*</span></div>)',
                r'\1' + bloque_prom,
                nuevo_html
            )
        # Inyectar PLUSV_SEMANA
        nuevo_html = re.sub(
            r'var PLUSV_SEMANA=[0-9.]+;',
            f'var PLUSV_SEMANA={mensual_data["plusv_semana"]};',
            nuevo_html
        )
        if 'var PLUSV_SEMANA=' not in nuevo_html:
            nuevo_html = nuevo_html.replace(
                'var RENDIMIENTO_MES=',
                f'var PLUSV_HOY={mensual_data["plusv_hoy"]};var PLUSV_SEMANA={mensual_data["plusv_semana"]};var RENDIMIENTO_MES='
            )
        nuevo_html = re.sub(r'var PLUSV_HOY=[0-9.]+;', f'var PLUSV_HOY={mensual_data["plusv_hoy"]};', nuevo_html)


    if ganancias_data:
        def fmtg(v):
            color = "var(--green)" if v >= 0 else "var(--red)"
            signo = "+" if v >= 0 else ""
            num = f"{abs(v):,.2f}".replace(",","X").replace(".",",").replace("X",".")
            return f'<span style="color:{color};font-size:20px;font-weight:500">{signo}{num} \u20ac</span>'
        def bloque(titulo, bruto, neto, id_b=None, id_n=None):
            ib = f' id="{id_b}"' if id_b else ''
            ib_n = f' id="{id_n}"' if id_n else ''
            return (
                f'<div style="padding:10px 0;border-bottom:1px solid var(--border)">' +
                f'<div style="font-size:11px;font-weight:700;color:var(--muted);letter-spacing:1px;margin-bottom:6px">{titulo}</div>' +
                f'<div style="display:flex;justify-content:space-between;align-items:baseline;margin-bottom:2px">' +
                f'<span style="font-size:12px;color:var(--muted)">bruto</span><span{ib} style="font-size:20px;font-weight:500">{fmtg(bruto)}</span></div>' +
                f'<div style="display:flex;justify-content:space-between;align-items:baseline">' +
                f'<span style="font-size:12px;color:var(--muted)">neto</span><span{ib_n} style="font-size:20px;font-weight:500">{fmtg(neto)}</span></div>' +
                f'</div>'
            )
        card = (
            '<div class="card" id="card-ganancias">' +
            '<div class="card-label">Cartera · Valores</div>' +
            bloque("HOY", 0, 0, id_b="hoy-b", id_n="hoy-n") +
            bloque("ESTA SEMANA", ganancias_data["sem_b"], ganancias_data["sem_n"], id_b="sem-b", id_n="sem-n") +
            bloque("ESTE MES", ganancias_data["mes_b"], ganancias_data["mes_n"], id_b="mes-b", id_n="mes-n") +
            bloque("ANUAL", ganancias_data["ani_b"], ganancias_data["ani_n"], id_b="anual-b", id_n="anual-n") +
            '</div>'
        )
        END_MARKER = '<!--/card-ganancias-->'
        card_con_marker = card + END_MARKER
        if 'id="card-ganancias"' in nuevo_html:
            start = nuevo_html.find('<div class="card" id="card-ganancias">')
            end = nuevo_html.find(END_MARKER)
            if end != -1:
                nuevo_html = nuevo_html[:start] + card_con_marker + nuevo_html[end+len(END_MARKER):]
            else:
                nuevo_html = nuevo_html[:start] + card_con_marker + nuevo_html[start:]
                nuevo_html = nuevo_html.replace(card_con_marker + card_con_marker, card_con_marker)
        else:
            nuevo_html = nuevo_html.replace('<p class="note">', card_con_marker + '\n  <p class="note">')
    if rendimiento_data:
        # Inyectar como variables JS
        js_vars = f'<script>var RENDIMIENTO_MES={rendimiento_data["mes"]};var RENDIMIENTO_MES_N={rendimiento_data["mes_n"]};var RENDIMIENTO_ANUAL={rendimiento_data["anual"]};var RENDIMIENTO_ANUAL_N={rendimiento_data["anual_n"]};</script>'
        if 'var RENDIMIENTO_MES=' in nuevo_html:
            import re as re4
            nuevo_html = re4.sub(r'<script>var RENDIMIENTO_MES=.*?;</script>', js_vars, nuevo_html)
        else:
            nuevo_html = nuevo_html.replace('</body>', js_vars + '</body>')
    if False and rendimiento_data:
        def fmtr(v):
            color = "var(--green)" if v >= 0 else "var(--red)"
            signo = "+" if v >= 0 else ""
            num = f"{abs(v):,.2f}".replace(",","X").replace(".",",").replace("X",".")
            return f'<span style="color:{color};font-size:20px;font-weight:500">{signo}{num} \u20ac</span>'
        import re as re3
        nuevo_html = re3.sub(
            r'ESTE MES</div>.*?bruto</span><span[^>]*>[^<]*</span>',
            lambda m: m.group(0)[:m.group(0).rfind('<span')] + fmtr(rendimiento_data["mes"]),
            nuevo_html, flags=re3.DOTALL, count=1
        )
        nuevo_html = re3.sub(
            r'ANUAL</div>.*?bruto</span><span[^>]*>[^<]*</span>',
            lambda m: m.group(0)[:m.group(0).rfind('<span')] + fmtr(rendimiento_data["anual"]),
            nuevo_html, flags=re3.DOTALL, count=1
        )
    if ventas_data is not None:
        nuevo_html = inyectar_ventas_anual(nuevo_html, ventas_data)
    nuevo_html = asegurar_proximas_compras(nuevo_html, proximas_compras or [])
    nuevo_html = asegurar_historico(nuevo_html)
    INDEX_HTML.write_text(nuevo_html, encoding="utf-8")
    print(f"✅ index.html actualizado. Backup: {backup.name}")



def actualizar_earnings_local():
    import os
    # Cargar .env si existe
    env_file = PROYECTO / '.env'
    if env_file.exists():
        for line in env_file.read_text().splitlines():
            if '=' in line and not line.startswith('#'):
                k, v = line.split('=', 1)
                os.environ.setdefault(k.strip(), v.strip())
    import json, csv, io
    from urllib.request import Request, urlopen
    from datetime import datetime
    from supabase import create_client
    sb_url = os.environ.get('SUPABASE_URL', 'https://ntvupakoulwiffvdcfox.supabase.co')
    sb_key = os.environ.get('SUPABASE_KEY', '')
    if not sb_key:
        print('⚠️  Sin SUPABASE_KEY, earnings omitidos'); return
    sb = create_client(sb_url, sb_key)
    tickers_data = json.loads((PROYECTO / 'tickers.json').read_text())
    sufijos = ['.MC','.AS','.DE','.PA','.MI','.BR','.LS']
    symbols = set()
    for s in tickers_data.keys():
        t = s.upper()
        for sf in sufijos: t = t.replace(sf,'')
        symbols.add(t)
    url = 'https://www.alphavantage.co/query?function=EARNINGS_CALENDAR&horizon=3month&apikey=9JNVRNFZ3D5VBP5E'
    data = urlopen(Request(url, headers={'User-Agent':'Mozilla/5.0'})).read().decode()
    reader = csv.DictReader(io.StringIO(data))
    hoy = datetime.now().strftime('%Y-%m-%d')
    count = 0
    for row in reader:
        sym = row['symbol'].strip().upper()
        if sym not in symbols: continue
        fecha = row['reportDate'].strip()
        if fecha < hoy: continue
        est = row.get('estimate','').strip()
        momento = row.get('timeOfTheDay','').strip()
        sb.table('earnings').upsert({'symbol':sym,'nombre':row.get('name','').strip(),'fecha':fecha,'estimacion':float(est) if est else None,'momento':momento or None},on_conflict='symbol,fecha').execute()
        count += 1
    print(f'✅ Earnings actualizados: {count} valores')

def verificar_precios_yahoo(ticker_map):
    """
    Verifica en tiempo real que todos los simbolos Yahoo devuelven precio.
    Llama al mismo endpoint que usa el servidor (v8/finance/chart).
    Si alguno falla, lo muestra en rojo con la solución exacta.
    """
    import urllib.request, urllib.parse, json as _json, time

    simbolos = [(tckr, info['yahoo']) for tckr, info in ticker_map.items()]
    total = len(simbolos)
    fallos = []

    print(f"\n🔍 Verificando precios de {total} tickers en Yahoo Finance...")

    BATCH = 10
    for i in range(0, len(simbolos), BATCH):
        batch = simbolos[i:i+BATCH]
        for tckr, sym in batch:
            url = f"https://query1.finance.yahoo.com/v8/finance/chart/{urllib.parse.quote(sym)}?interval=1d&range=1d"
            try:
                req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
                with urllib.request.urlopen(req, timeout=6) as r:
                    data = _json.loads(r.read().decode())
                result = data.get("chart", {}).get("result")
                precio = result[0]["meta"].get("regularMarketPrice") if result else None
                if not precio:
                    fallos.append((tckr, sym))
            except Exception:
                fallos.append((tckr, sym))
        time.sleep(0.3)  # respetar rate limit Yahoo

    ok = total - len(fallos)
    if not fallos:
        print(f"✅ {ok}/{total} tickers con precio OK")
    else:
        print(f"\n{'='*50}")
        print(f"⚠️  {ok}/{total} tickers con precio — {len(fallos)} SIN PRECIO:")
        for tckr, sym in fallos:
            moneda = ticker_map[tckr].get('moneda', '?')
            fuente = ticker_map[tckr].get('fuente_resolucion', '?')
            print(f"\n  ❌ {tckr} -> '{sym}' no devuelve precio en Yahoo")
            if fuente == 'fallback_madrid':
                print(f"     ⚡ ACCION: añade en tickers_override.json:")
                print(f'     \033[91m"{tckr}": "SIMBOLO_CORRECTO"\033[0m')
                print(f"     (busca el símbolo en finance.yahoo.com)")
            elif fuente == 'fallback_usd_directo':
                print(f"     ⚡ ACCION: añade en tickers_override.json:")
                print(f'     \033[91m"{tckr}": "SIMBOLO_CORRECTO"\033[0m')
            else:
                print(f"     ⚡ ACCION: verifica el símbolo en finance.yahoo.com")
        print(f"{'='*50}")
        print(f"\033[91m⛔ ATENCION: estos tickers no se mostraran en la app hasta corregirlos\033[0m")


def leer_ventas_anual():
    """Lee ventas del año actual del Excel (pestaña 2026, filas 262-400)."""
    from datetime import datetime
    wb2 = openpyxl.load_workbook(open(str(EXCEL), "rb"), read_only=True, data_only=True)
    ws2 = wb2["2026"]
    hoy = datetime.now()
    ventas = []
    for row in ws2.iter_rows(min_row=262, max_row=400, values_only=True):
        ticker = row[3]   # col D
        moneda = row[6]   # col G
        banco  = row[7]   # col H
        fecha  = row[16]  # col Q
        bruto  = row[24]  # col Y
        neto   = row[25]  # col Z
        if not ticker or not fecha or bruto is None or neto is None:
            continue
        if not isinstance(fecha, datetime):
            continue
        if fecha.year != hoy.year:
            continue
        ventas.append({
            "ticker": str(ticker).strip(),
            "fecha":  fecha.strftime("%Y-%m-%d"),
            "bruto":  round(float(bruto), 2),
            "neto":   round(float(neto),  2),
            "banco":  normalizar_banco(banco),
            "moneda": str(moneda).strip() if moneda else "EUR",
        })
    print(f"\u2705 {len(ventas)} ventas leidas del Excel (VENTAS_ANUAL)")
    return ventas


def inyectar_ventas_anual(html, ventas):
    """Reemplaza var VENTAS_ANUAL en index.html con marcadores robustos.
    También elimina copias duplicadas de window.renderVentas."""
    import re as _re
    import json as _json

    ventas_json = _json.dumps(ventas, ensure_ascii=False, separators=(",", ":"))
    bloque = (
        "<!--ventas-anual-start-->"
        f"<script>var VENTAS_ANUAL={ventas_json};</script>"
        "<!--ventas-anual-end-->"
    )

    # Sustituir si ya hay marcadores
    if "<!--ventas-anual-start-->" in html:
        html = _re.sub(
            r"<!--ventas-anual-start-->.*?<!--ventas-anual-end-->",
            bloque, html, flags=_re.DOTALL
        )
    # Sustituir el script sin marcadores (primera vez)
    elif "<script>var VENTAS_ANUAL=[" in html:
        html = _re.sub(
            r"<script>var VENTAS_ANUAL=\[.*?\];</script>",
            bloque, html, flags=_re.DOTALL
        )
    else:
        html = html.replace("</body>", bloque + "\n</body>")

    # ── Eliminar renderVentas duplicados: conservar solo el primero ────────
    patron = r"window\.renderVentas\s*=\s*function\s*\(\)"
    ocurrencias = [m.start() for m in _re.finditer(patron, html)]
    if len(ocurrencias) > 1:
        # Localizar fin de la primera definición contando llaves
        inicio = ocurrencias[0]
        depth = 0
        fin = -1
        for i in range(inicio, len(html)):
            if html[i] == "{": depth += 1
            elif html[i] == "}":
                depth -= 1
                if depth == 0: fin = i; break
        if fin != -1:
            primera = html[:fin + 1]
            resto = html[fin + 1:]
            resto_limpio = _re.sub(
                r"window\.renderVentas\s*=\s*function\s*\(\)\s*\{.*?\}\s*;",
                "", resto, flags=_re.DOTALL
            )
            html = primera + resto_limpio
            n_eliminadas = len(ocurrencias) - 1
            print(f"\u2705 renderVentas deduplicado: {n_eliminadas} copia(s) extra eliminada(s)")

    return html


def main():
    print(f"📂 Leyendo Excel: {EXCEL}")
    posiciones, sin_mic, mensual_data, compras_por_ticker, proximas_compras = leer_excel_con_mic()
    ganancias_data = leer_ganancias_realizadas()
    rendimiento_data = leer_rendimiento_cartera()
    print(f"✅ {len(posiciones)} tickers unicos cargados")

    overrides = cargar_overrides()
    print(f"✅ {len(overrides)} overrides manuales")

    # Resolver simbolos
    ticker_map = {}
    stats = {}
    nuevos_overrides = {}

    for p in posiciones:
        sym, fuente = resolver_simbolo_yahoo(p, overrides)

        # Si no se pudo resolver -> buscar automaticamente en Yahoo Finance
        if fuente == 'sin_resolver':
            print(f"   🔍 Buscando {p['tckr']} en Yahoo Finance...")
            sym_auto = buscar_simbolo_yahoo_auto(p['tckr'])
            if sym_auto:
                sym = sym_auto
                fuente = 'auto_yahoo'
                nuevos_overrides[p['tckr']] = sym_auto
                print(f"   ✅ {p['tckr']} -> {sym_auto} (guardado en overrides)")
            else:
                print(f"   ❌ {p['tckr']} no encontrado en Yahoo — verifica manualmente")

        ticker_map[p["tckr"]] = {
            "yahoo": sym,
            "mic": p.get("mic"),
            "moneda": p["moneda"],
            "titulos": float(p["titulos"]),
            "coste_eur": round(p["coste_eur"], 4),
            "precio_excel": p["precio_excel"] if isinstance(p["precio_excel"], (int, float)) else None,
            "fuente_resolucion": fuente,
        }
        stats[fuente] = stats.get(fuente, 0) + 1

    # Guardar nuevos overrides encontrados automaticamente
    if nuevos_overrides:
        for ticker, symbol in nuevos_overrides.items():
            guardar_override(ticker, symbol)
        print(f"\n✅ {len(nuevos_overrides)} nuevos overrides guardados automaticamente en tickers_override.json")

    print(f"\n📊 Resolucion de simbolos:")
    for fuente, n in sorted(stats.items(), key=lambda x: -x[1]):
        print(f"   {fuente:<25} {n}")

    if sin_mic:
        print(f"\n⚠️  {len(sin_mic)} tickers SIN MIC en Excel:")
        for t in sin_mic[:20]:
            print(f"     - {t}")
        if len(sin_mic) > 20:
            print(f"     ... y {len(sin_mic)-20} mas")
        print(f"   Estos usaran fallback (USD directo o .MC). Si descuadran, anade")
        print(f"   liena tipo 'Nombre (MIC:TICKER)' a {MIC_NAMES_TXT.name}")

    with open(TICKERS_JSON, "w", encoding="utf-8") as f:
        json.dump(ticker_map, f, indent=2, ensure_ascii=False)
    print(f"\n✅ Guardado {TICKERS_JSON}")

    const_C = construir_const_C_compacta(posiciones, ticker_map, compras_por_ticker)
    ventas_data = leer_ventas_anual()
    actualizar_index_html(const_C, mensual_data, ganancias_data=ganancias_data, rendimiento_data=rendimiento_data, proximas_compras=proximas_compras, ventas_data=ventas_data)

    # ══════════════════════════════════════════════════
    # VERIFICACION DE PRECIOS — avisa si algun ticker
    # no devuelve precio desde Yahoo antes del push
    # ══════════════════════════════════════════════════
    verificar_precios_yahoo(ticker_map)

    print("\n🎯 LISTO. Siguiente paso:")
    print("   python3 validate.py")
    try:
        actualizar_earnings_local()
    except Exception as e:
        print(f"⚠️  Earnings: {e}")


if __name__ == "__main__":
    main()
# version updater - añadido automaticamente

